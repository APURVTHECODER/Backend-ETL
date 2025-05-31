# backend/routers/export.py

import json
import os
import logging
import re
import io
import traceback # Ensure traceback is imported for detailed error logging
from typing import List, Dict, Any, Optional, Set
import uuid
import pandas as pd
from dotenv import load_dotenv
from fastapi import APIRouter, Depends, HTTPException, Response, status # Import status
from pydantic import BaseModel, Field
from google.cloud import bigquery # Keep for type hints
from google.api_core.exceptions import GoogleAPICallError, NotFound, Forbidden
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, date, time, timezone # Added timezone
import base64
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font

# --- Import Config Variables ---
# Import only specific variables needed from config
from config import DEFAULT_BQ_LOCATION, DEFAULT_JOB_TIMEOUT_SECONDS
from dateutil.parser import parse as dateutil_parse 
# --- Import Dependencies ---
from auth import verify_token # For authentication
from dependencies.client_deps import get_bigquery_client # Import client dependency getter

# --- Get Logger for this module ---
logger_export = logging.getLogger(__name__) # Get a logger specific to this module

load_dotenv() # Load .env if used for local development config

# --- Configuration ---
MAX_SOURCE_TABLE_ROWS = int(os.getenv("EXPORT_MAX_SOURCE_ROWS", 5000)) # Example: make configurable

# --- Router Setup ---
export_router = APIRouter(
    prefix="/api/export",
    tags=["Export"],
    dependencies=[Depends(verify_token)] # Apply authentication to all export routes
)
class ChartConfig(BaseModel):
    type: str
    x_axis: str
    y_axes: List[str]
    rationale: Optional[str] = None

class QueryToExcelRequest(BaseModel):
    job_id: str
    sql: str
    location: str
    chart_image_base64: Optional[str] = Field(None) # No '...' default for Optional
    chart_config: Optional[ChartConfig] = Field(None)
    chart_data: Optional[List[Dict[str, Any]]] = Field(None)
    data_override: Optional[List[Dict[str, Any]]] = Field(None, description="If provided, use this data instead of re-fetching full results.")
    schema_override: Optional[List[Dict[str, Any]]] = Field(None, description="Schema for the data_override. Each dict: {'name': str, 'type': str}")
# --- Pydantic Models ---
class QueryExportRequest(BaseModel):
    sql: str = Field(..., description="The SQL query that was executed.")
    job_id: str = Field(..., description="The BigQuery Job ID for the executed query.")
    location: Optional[str] = Field(DEFAULT_BQ_LOCATION, description="The location where the job ran.")

# --- Helper Functions ---

def extract_fully_qualified_tables(sql: str) -> List[str]:
    """
    Extracts fully qualified table names (project.dataset.table) from SQL.
    Handles backticks and preserves original casing.
    """
    # Regex to find fully qualified tables (project.dataset.table or dataset.table) after FROM or JOIN
    # Handles optional backticks around the full name or individual parts
    regex = r"""
        (?:FROM|JOIN)\s+             # Match FROM or JOIN followed by space
        `?                           # Optional starting backtick for the whole name
        (                            # Start capturing group 1 (full qualified name)
          (?:                        # Start non-capturing group for project/dataset parts
            `?                       # Optional backtick for part
            [a-zA-Z0-9_.-]+          # Match project/dataset name characters
            `?                       # Optional closing backtick for part
            \.                       # Match the dot separator
          )+                         # Match one or more project/dataset parts (e.g., proj.dataset.)
          `?                         # Optional backtick for table name part
          [a-zA-Z0-9_-]+             # Match table name characters
          `?                         # Optional closing backtick for table name part
        )                            # End capturing group 1
        `?                           # Optional closing backtick for the whole name
    """
    matches = re.finditer(regex, sql, re.VERBOSE | re.IGNORECASE | re.MULTILINE)
    tables_dict: Dict[str, str] = {} # Use dict to store unique tables (lowercase key, original case value)
    for match in matches:
        table_name = match.group(1).replace('`', '') # Remove all backticks
        # Ensure it looks like at least dataset.table
        if table_name.count('.') >= 1:
            lower_case_key = table_name.lower()
            # Store the first occurrence with original casing
            if lower_case_key not in tables_dict:
                tables_dict[lower_case_key] = table_name
    original_case_tables = list(tables_dict.values())
    logger_export.info(f"Extracted source tables (preserved case): {original_case_tables}")
    return original_case_tables

# --- MODIFIED HELPER to accept bq_client ---
def fetch_bq_data_to_dataframe(
    query: str,
    location: Optional[str], # Location might be needed for the query job config
    bq_client: bigquery.Client # Accept the initialized client
) -> pd.DataFrame:
    """
    Fetches data using a query and returns a Pandas DataFrame using the provided client.
    Handles common errors and timeouts.
    """
    query_job = None # Define here for access in exception blocks
    # Use the timeout defined in config, defaulting if not set
    timeout_seconds = DEFAULT_JOB_TIMEOUT_SECONDS

    try:
        logger_export.info(f"Executing BQ query for export helper: {query[:150]}...")
        # Use the passed BigQuery client instance
        query_job_config = bigquery.QueryJobConfig() # Add job config if needed (e.g., query parameters)
        query_job = bq_client.query(query, location=location, job_config=query_job_config)
        logger_export.debug(f"Waiting for helper query job {query_job.job_id} (timeout: {timeout_seconds}s)...")

        # Use to_dataframe with timeout - this waits for the job to complete
        df = query_job.to_dataframe()

        logger_export.info(f"Helper fetched {len(df)} rows into DataFrame for job {query_job.job_id}.")
        return df

    except TimeoutError: # Catch specific timeout from to_dataframe(timeout=...)
        job_id_str = getattr(query_job, 'job_id', 'unknown')
        logger_export.error(f"Helper query job {job_id_str} timed out after {timeout_seconds} seconds.")
        # Return a DataFrame indicating the error for this specific source table
        return pd.DataFrame([{"error": f"Source table query timed out ({timeout_seconds}s)", "query": query}])
    except NotFound as e:
        logger_export.warning(f"Helper query error (NotFound): {e}. Query: {query[:150]}")
        return pd.DataFrame([{"error": f"Source table/resource not found: {e.message}", "query": query}])
    except Forbidden as e:
         logger_export.warning(f"Helper query error (Forbidden): {e}. Query: {query[:150]}")
         return pd.DataFrame([{"error": f"Permission denied for source table: {e.message}", "query": query}])
    except (GoogleAPICallError, ValueError, TypeError) as e: # Catch other potential API or data errors
        logger_export.error(f"Helper query error: {e}. Query: {query[:150]}", exc_info=True)
        return pd.DataFrame([{"error": f"Failed source data fetch: {type(e).__name__} - {str(e)}", "query": query}])
    except Exception as e: # Catch any other unexpected errors
        logger_export.error(f"Unexpected error in helper query: {e}. Query: {query[:150]}", exc_info=True)
        return pd.DataFrame([{"error": f"Unexpected error fetching source: {str(e)}", "query": query}])


def make_datetime_naive(df: pd.DataFrame) -> pd.DataFrame:
    """Converts timezone-aware datetime columns in a DataFrame to timezone-naive for Excel."""
    if df is None or df.empty:
        return df
    # Select columns that are explicitly timezone-aware
    tz_aware_cols = df.select_dtypes(include=['datetime64[ns, UTC]', 'datetimetz']).columns
    if not tz_aware_cols.empty:
        logger_export.debug(f"Found timezone-aware columns: {list(tz_aware_cols)}")
        for col in tz_aware_cols:
            try:
                # Double-check with attribute access before converting
                if getattr(df[col].dt, 'tz', None) is not None:
                    logger_export.debug(f"Converting column '{col}' to timezone-naive...")
                    # tz_convert(None) is generally preferred over tz_localize(None)
                    # It converts to UTC first if necessary, then removes tz info.
                    df[col] = df[col].dt.tz_convert(None)
            except Exception as e:
                logger_export.warning(f"Could not convert column '{col}' to naive datetime: {e}. Skipping conversion.")
    return df


# --- API Endpoint (MODIFIED) ---
@export_router.post("/query-to-excel",
                    response_class=Response,
                    summary="Export Query Results, SQL, Source Previews, and Optional Chart to Excel",
                    description="Exports various data components related to a BigQuery job to an Excel file.",
                    responses={
                        200: {
                            "description": "Excel file generated successfully.",
                            "content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {
                                "schema": {"type": "string", "format": "binary"}
                            }}
                        },
                        # ... other responses
                    })




@export_router.post("/query-to-excel",
                    response_class=Response,
                    summary="Export Query Results, SQL, Source Previews, and Optional Chart to Excel",
                    # ... other decorator arguments
                   )
def export_query_to_excel(
    payload: QueryToExcelRequest,
    bq_client: bigquery.Client = Depends(get_bigquery_client)
):
    logger_export.info(f"Received Excel export request for Job ID: {payload.job_id}. Data override provided: {'Yes' if payload.data_override is not None else 'No'}. Chart included: {'Yes' if payload.chart_image_base64 else 'No'}")

    if not bq_client:
         logger_export.error("Export Endpoint: BigQuery client is None via dependency.")
         raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="BigQuery client dependency failed.")

    try:
        workbook = Workbook()
        results_sheet = workbook.active
        results_sheet.title = "Query Results"

        result_rows_for_excel: List[List[Any]] = []
        result_schema_names_for_excel: List[str] = []

        # --- +++ LOGIC TO USE DATA_OVERRIDE OR FALLBACK +++ ---
        if payload.data_override is not None:
            if len(payload.data_override) > 0:
                logger_export.info(f"Using data_override for 'Query Results' sheet. Rows: {len(payload.data_override)}")
                
                # Determine schema: use schema_override if provided, else infer from data_override keys
                if payload.schema_override and len(payload.schema_override) > 0:
                    result_schema_names_for_excel = [field.get("name", f"UnknownCol_{i}") for i, field in enumerate(payload.schema_override)]
                    logger_export.info(f"Using schema_override for headers: {result_schema_names_for_excel}")
                else: # Infer schema from first row of data_override
                    result_schema_names_for_excel = list(payload.data_override[0].keys())
                    logger_export.info(f"Inferred schema from data_override keys: {result_schema_names_for_excel}")
                
                # Convert list of dicts (data_override) to list of lists for openpyxl
                for row_dict in payload.data_override:
                    row_values = []
                    for col_name in result_schema_names_for_excel:
                        value = row_dict.get(col_name)
                        # Apply serialization for consistency
                        if isinstance(value, str) and re.match(r'^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}', value):
                            try:
                                # Parse ISO string from frontend back into a datetime object
                                dt_obj = dateutil_parse(value)
                                if dt_obj.tzinfo is not None:
                                    dt_obj = dt_obj.astimezone(timezone.utc).replace(tzinfo=None)
                                value = dt_obj
                            except (ValueError, TypeError):
                                pass # Keep as string if parsing fails
                        elif isinstance(value, list) or isinstance(value, dict):
                            value = json.dumps(value, default=str)
                        row_values.append(value)
                    result_rows_for_excel.append(row_values)
            else: # data_override is an empty list
                logger_export.info("data_override was provided but is empty. Creating an empty 'Query Results' sheet with headers if possible.")
                if payload.schema_override and len(payload.schema_override) > 0:
                    result_schema_names_for_excel = [field.get("name", f"UnknownCol_{i}") for i, field in enumerate(payload.schema_override)]
                else: # No schema and no data, create a placeholder
                    result_schema_names_for_excel = ["Information"]
                    result_rows_for_excel.append(["Filtered data set was empty."])

        else: # Fallback to fetching full results from BigQuery if no data_override
            logger_export.info(f"No data_override. Fetching full results for job {payload.job_id} from BigQuery.")
            try:
                job = bq_client.get_job(payload.job_id, location=payload.location)
            except NotFound:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Job ID '{payload.job_id}' not found.")

            if job.state != 'DONE': raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Job {payload.job_id} not complete.")
            if job.error_result: raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Job {payload.job_id} failed: {job.error_result.get('message', 'Unknown')}")

            if not job.destination:
                result_schema_names_for_excel = ["Information"]
                result_rows_for_excel.append([f"Query (Job ID: {job.job_id}) did not produce a result table."])
                if job.statement_type: result_rows_for_excel.append([f"Statement Type: {job.statement_type}"])
                if hasattr(job, 'num_dml_affected_rows') and job.num_dml_affected_rows is not None: result_rows_for_excel.append([f"Rows Affected: {job.num_dml_affected_rows}"])
            else:
                rows_iterator = bq_client.list_rows(job.destination, timeout=DEFAULT_JOB_TIMEOUT_SECONDS)
                if rows_iterator.schema:
                    result_schema_names_for_excel = [field.name for field in rows_iterator.schema]
                
                for row_data in rows_iterator:
                    serialized_values = []
                    for field_name in result_schema_names_for_excel:
                        value = row_data[field_name]
                        if isinstance(value, bytes):
                            try: serialized_values.append(value.decode('utf-8'))
                            except UnicodeDecodeError: serialized_values.append(f"0x{value.hex()}")
                        elif isinstance(value, (datetime, pd.Timestamp)):
                            if value.tzinfo is not None: value = value.astimezone(timezone.utc).replace(tzinfo=None)
                            serialized_values.append(value)
                        elif isinstance(value, date): serialized_values.append(value)
                        elif isinstance(value, time): serialized_values.append(value.isoformat())
                        elif isinstance(value, list) or isinstance(value, dict): serialized_values.append(json.dumps(value, default=str))
                        else: serialized_values.append(value)
                    result_rows_for_excel.append(serialized_values)
                logger_export.info(f"Fallback: Fetched {len(result_rows_for_excel)} result rows for job {payload.job_id}.")

        # --- Write headers and rows for "Query Results" sheet ---
        if result_schema_names_for_excel:
            results_sheet.append(result_schema_names_for_excel)
        for row_values in result_rows_for_excel:
            results_sheet.append(row_values)
        logger_export.info(f"Main query data written to Excel sheet 'Query Results'.")

        # --- Sheet 2: Original SQL Query (No changes) ---
        query_sheet = workbook.create_sheet(title="SQL Query")
        query_sheet.cell(row=1, column=1, value="Executed SQL Query:").font = Font(bold=True)
        query_sheet.cell(row=2, column=1, value=payload.sql)
        query_sheet.column_dimensions['A'].width = 80
        logger_export.info("SQL query written to Excel sheet 'SQL Query'.")

        # --- Sheets 3+: Source Table Previews (No changes) ---
        source_tables = extract_fully_qualified_tables(payload.sql)
        processed_sheet_names: Set[str] = set(["Query Results", "SQL Query"])

        for table_fqn in source_tables:
            # ... (your existing source table preview logic - no changes needed) ...
            if 'information_schema' in table_fqn.lower(): continue
            logger_export.info(f"Fetching preview for source table: {table_fqn} (Limit: {MAX_SOURCE_TABLE_ROWS})")
            preview_query = f"SELECT * FROM `{table_fqn}` LIMIT {MAX_SOURCE_TABLE_ROWS}"
            df_source_preview = fetch_bq_data_to_dataframe(preview_query, payload.location, bq_client)
            df_source_preview_naive = make_datetime_naive(df_source_preview.copy())
            base_table_name = table_fqn.split('.')[-1]; sanitized_base_name = re.sub(r'[\\/*?:\[\]]', '_', base_table_name)[:25]
            sheet_title_prefix = "Source_"
            if 'error' in df_source_preview_naive.columns and len(df_source_preview_naive) == 1: sheet_title_prefix = "Error_Source_"
            sheet_name_candidate = f"{sheet_title_prefix}{sanitized_base_name}"; final_sheet_name = sheet_name_candidate; count = 1
            while final_sheet_name.lower()[:31] in (name.lower()[:31] for name in processed_sheet_names):
                suffix = f"_{count}"; final_sheet_name = f"{sheet_name_candidate[:31-len(suffix)]}{suffix}"; count += 1
                if count > 99: final_sheet_name = f"{sheet_title_prefix}Table_{uuid.uuid4().hex[:8]}"; break
            final_sheet_name = final_sheet_name[:31]
            source_sheet = workbook.create_sheet(title=final_sheet_name)
            if df_source_preview_naive is not None and not df_source_preview_naive.empty:
                for r_idx, row_values_tuple in enumerate(dataframe_to_rows(df_source_preview_naive, index=False, header=True), 1):
                    for c_idx, value in enumerate(row_values_tuple, 1):
                        cell_to_write = "" if pd.isna(value) else value
                        try: source_sheet.cell(row=r_idx, column=c_idx, value=cell_to_write)
                        except Exception as cell_write_error:
                            logger_export.warning(f"Sheet '{final_sheet_name}': Error writing value {cell_to_write!r} to cell ({r_idx},{c_idx}). Error: {cell_write_error}.")
                            source_sheet.cell(row=r_idx, column=c_idx, value="ERROR_VAL")
            else: source_sheet.cell(row=1, column=1, value="No data or error fetching preview.")
            processed_sheet_names.add(final_sheet_name)
            logger_export.info(f"Source table preview '{table_fqn}' written to sheet '{final_sheet_name}'.")

        # --- Sheet N+1: Chart Image and Info (No changes) ---
        if payload.chart_image_base64 and payload.chart_config:
            # ... (your existing chart image and info logic) ...
            chart_title = f"{payload.chart_config.type.capitalize()} Chart"; unique_chart_sheet_title = chart_title; count = 1
            while unique_chart_sheet_title.lower()[:31] in (name.lower()[:31] for name in processed_sheet_names):
                suffix = f"_{count}"; unique_chart_sheet_title = f"{chart_title[:31-len(suffix)]}{suffix}"; count +=1
            unique_chart_sheet_title = unique_chart_sheet_title[:31]
            chart_info_sheet = workbook.create_sheet(title=unique_chart_sheet_title); processed_sheet_names.add(unique_chart_sheet_title)
            img_row_idx = 1
            chart_info_sheet.cell(row=img_row_idx, column=1, value="Chart Type:").font = Font(bold=True); chart_info_sheet.cell(row=img_row_idx, column=2, value=payload.chart_config.type); img_row_idx+=1
            chart_info_sheet.cell(row=img_row_idx, column=1, value="X-Axis:").font = Font(bold=True); chart_info_sheet.cell(row=img_row_idx, column=2, value=payload.chart_config.x_axis); img_row_idx+=1
            chart_info_sheet.cell(row=img_row_idx, column=1, value="Y-Axes:").font = Font(bold=True); chart_info_sheet.cell(row=img_row_idx, column=2, value=", ".join(payload.chart_config.y_axes)); img_row_idx+=1
            if payload.chart_config.rationale:
                 chart_info_sheet.cell(row=img_row_idx, column=1, value="Rationale:").font = Font(bold=True); chart_info_sheet.cell(row=img_row_idx, column=2, value=payload.chart_config.rationale); img_row_idx+=1
            img_row_idx+=1
            try:
                image_data = base64.b64decode(payload.chart_image_base64); image_stream = io.BytesIO(image_data); img = OpenpyxlImage(image_stream); chart_info_sheet.add_image(img, f"A{img_row_idx}")
            except Exception as img_e: logger_export.error(f"Failed to add image: {img_e}", exc_info=True)


        # --- Sheet N+2: Raw Chart Data (No changes needed, uses payload.chart_data) ---
        if payload.chart_data: # Use chart_data instead of chart_image_base64 as the condition
            # ... (your existing chart data sheet logic) ...
            # This part already correctly uses payload.chart_data which is the filtered data.
            chart_data_title = "Chart Source Data"; unique_chart_data_sheet_title = chart_data_title; count = 1
            while unique_chart_data_sheet_title.lower()[:31] in (name.lower()[:31] for name in processed_sheet_names):
                suffix = f"_{count}"; unique_chart_data_sheet_title = f"{chart_data_title[:31-len(suffix)]}{suffix}"; count +=1
            unique_chart_data_sheet_title = unique_chart_data_sheet_title[:31]
            chart_data_sheet = workbook.create_sheet(title=unique_chart_data_sheet_title); processed_sheet_names.add(unique_chart_data_sheet_title)
            df_chart = pd.DataFrame(payload.chart_data)
            df_chart_naive = make_datetime_naive(df_chart.copy())
            if not df_chart_naive.empty:
                for r_idx, row_val_list in enumerate(dataframe_to_rows(df_chart_naive, index=False, header=True), 1):
                    for c_idx, value in enumerate(row_val_list, 1):
                        cell_to_write = "" if pd.isna(value) else value
                        try: chart_data_sheet.cell(row=r_idx, column=c_idx, value=cell_to_write)
                        except Exception: chart_data_sheet.cell(row=r_idx, column=c_idx, value="ERROR_VAL")
            logger_export.info(f"Chart source data ({len(df_chart_naive)} rows) added to '{unique_chart_data_sheet_title}' sheet.")


        # --- Save workbook to a BytesIO stream ---
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S") # Use UTC
        filename = f"query_export_{payload.job_id[:8]}_{timestamp}.xlsx"
        
        headers = { 'Content-Disposition': f'attachment; filename="{filename}"' }
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )

    except HTTPException as http_exc:
        raise http_exc
    except GoogleAPICallError as e:
        logger_export.error(f"Google API Error during export for job {payload.job_id}: {e}", exc_info=True)
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=f"Error communicating with Google Cloud: {str(e)}")
    except Exception as e:
        logger_export.error(f"Unexpected error during Excel export for job {payload.job_id}: {e}", exc_info=True)
        logger_export.error(traceback.format_exc())
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Failed to generate export file: {str(e)}")